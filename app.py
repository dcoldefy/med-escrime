import io
import json
import os
import urllib.request as _urllib_req
import uuid
from fastapi import FastAPI, Depends, Request, HTTPException, UploadFile, File, Form, Header
from fastapi.responses import Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
from datetime import date as dt_date, time as dt_time, datetime

import database

app = FastAPI(title="MED - My Escrime Data")

UPLOAD_DIR = "static/uploads"

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


@app.on_event("startup")
def startup_event():
    database.init_db()
    os.makedirs(UPLOAD_DIR, exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
#  Auth
# ─────────────────────────────────────────────────────────────────────────────

def get_current_user(
    x_user_token: Optional[str] = Header(None),
    db: Session = Depends(database.get_db),
) -> database.User:
    if not x_user_token:
        raise HTTPException(status_code=403, detail="Token manquant")
    user = db.query(database.User).filter(database.User.token == x_user_token).first()
    if not user:
        raise HTTPException(status_code=403, detail="Token invalide")
    return user


def require_admin(request: Request):
    """Admin accessible uniquement via Tailscale ou réseau local (pas via Cloudflare)."""
    if request.headers.get("CF-Connecting-IP"):
        raise HTTPException(status_code=403, detail="Admin non accessible depuis internet")
    host = request.client.host if request.client else ""
    if not (host.startswith("100.") or host in ("127.0.0.1", "::1")):
        raise HTTPException(status_code=403, detail="Accès refusé")


# ─────────────────────────────────────────────────────────────────────────────
#  Pages
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/")
def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/api/me")
def get_me(user: database.User = Depends(get_current_user)):
    return {"id": user.id, "name": user.name}


@app.get("/admin")
def admin_page(request: Request, db: Session = Depends(database.get_db)):
    require_admin(request)
    users = db.query(database.User).order_by(database.User.created_at).all()
    return templates.TemplateResponse("admin.html", {"request": request, "users": users})


@app.get("/admin/data")
def data_page(request: Request, db: Session = Depends(database.get_db)):
    require_admin(request)
    users = db.query(database.User).order_by(database.User.created_at).all()
    return templates.TemplateResponse("data.html", {"request": request, "users": users})


# ─────────────────────────────────────────────────────────────────────────────
#  Admin — gestion utilisateurs
# ─────────────────────────────────────────────────────────────────────────────

class UserPayload(BaseModel):
    name: str


@app.post("/api/admin/users")
def create_user(request: Request, payload: UserPayload, db: Session = Depends(database.get_db)):
    require_admin(request)
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nom obligatoire")
    user = database.User(name=name, token=str(uuid.uuid4()))
    db.add(user)
    db.commit()
    db.refresh(user)
    return _user_dict(user)


@app.get("/api/admin/users")
def list_users(request: Request, db: Session = Depends(database.get_db)):
    require_admin(request)
    users = db.query(database.User).order_by(database.User.created_at).all()
    return [_user_dict(u) for u in users]


@app.get("/api/admin/users/{user_id}/competitions")
def admin_list_competitions(user_id: int, request: Request, db: Session = Depends(database.get_db)):
    require_admin(request)
    rows = (
        db.query(database.Competition)
        .filter(database.Competition.user_id == user_id)
        .order_by(database.Competition.date.desc())
        .all()
    )
    return [_competition_dict(c) for c in rows]


@app.get("/api/admin/users/{user_id}/assaults")
def admin_list_assaults(user_id: int, request: Request, db: Session = Depends(database.get_db)):
    require_admin(request)
    rows = (
        db.query(database.Assault)
        .filter(database.Assault.user_id == user_id)
        .order_by(database.Assault.date.desc())
        .all()
    )
    return [_assault_dict(a) for a in rows]


@app.get("/api/admin/competitions/{comp_id}/detail")
def admin_competition_detail(comp_id: int, request: Request, db: Session = Depends(database.get_db)):
    require_admin(request)
    c = _get_or_404(db, database.Competition, comp_id)
    result = _competition_dict(c)

    p = db.query(database.Poule).filter(database.Poule.competition_id == comp_id).first()
    if p:
        assaults_poule = (
            db.query(database.AssaultPoule)
            .filter(database.AssaultPoule.poule_id == p.id)
            .order_by(database.AssaultPoule.numero)
            .all()
        )
        result["poule"] = {**_poule_dict(p), "assaults": [_assault_poule_dict(a) for a in assaults_poule]}
    else:
        result["poule"] = None

    tableau = (
        db.query(database.AssaultTableau)
        .filter(database.AssaultTableau.competition_id == comp_id)
        .order_by(database.AssaultTableau.tour)
        .all()
    )
    result["tableau"] = [_assault_tableau_dict(a) for a in tableau]
    return result


@app.delete("/api/admin/users/{user_id}")
def delete_user(user_id: int, request: Request, db: Session = Depends(database.get_db)):
    require_admin(request)
    user = db.query(database.User).filter(database.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")
    db.delete(user)
    db.commit()
    return {"ok": True}


# ─────────────────────────────────────────────────────────────────────────────
#  Entraînement — assaults libres
# ─────────────────────────────────────────────────────────────────────────────

class AssaultPayload(BaseModel):
    date: str
    heure: str
    notes: Optional[str] = ""


@app.post("/api/assaults")
def create_assault(
    payload: AssaultPayload,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    a = database.Assault(
        user_id=user.id,
        date=dt_date.fromisoformat(payload.date),
        heure=dt_time.fromisoformat(payload.heure),
        type_seance="entrainement",
        notes=payload.notes.strip() if payload.notes else "",
    )
    db.add(a)
    db.commit()
    db.refresh(a)
    return _assault_dict(a)


@app.get("/api/assaults")
def list_assaults(
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    rows = (
        db.query(database.Assault)
        .filter(database.Assault.user_id == user.id)
        .order_by(database.Assault.date.desc(), database.Assault.heure.desc())
        .all()
    )
    return [_assault_dict(a) for a in rows]


@app.get("/api/assaults/{assault_id}")
def get_assault(
    assault_id: int,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    a = _get_or_404(db, database.Assault, assault_id)
    _check_owner(a.user_id, user.id)
    return _assault_dict(a)


@app.patch("/api/assaults/{assault_id}")
def patch_assault(
    assault_id: int,
    data: dict,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    a = _get_or_404(db, database.Assault, assault_id)
    _check_owner(a.user_id, user.id)
    for k, v in data.items():
        if k == "date":
            a.date = dt_date.fromisoformat(v)
        elif k == "heure":
            a.heure = dt_time.fromisoformat(v)
        elif k == "notes":
            a.notes = v
    db.commit()
    db.refresh(a)
    return _assault_dict(a)


@app.delete("/api/assaults/{assault_id}")
def delete_assault(
    assault_id: int,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    a = _get_or_404(db, database.Assault, assault_id)
    _check_owner(a.user_id, user.id)
    db.delete(a)
    db.commit()
    return {"ok": True}


# ─────────────────────────────────────────────────────────────────────────────
#  Leçons
# ─────────────────────────────────────────────────────────────────────────────

class LeconPayload(BaseModel):
    date: str
    heure: str
    maitre: Optional[str] = ""
    theme: Optional[str] = ""
    notes: Optional[str] = ""


@app.post("/api/lecons")
def create_lecon(
    payload: LeconPayload,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    l = database.Lecon(
        user_id=user.id,
        date=dt_date.fromisoformat(payload.date),
        heure=dt_time.fromisoformat(payload.heure),
        maitre=payload.maitre.strip() if payload.maitre else "",
        theme=payload.theme.strip() if payload.theme else "",
        notes=payload.notes.strip() if payload.notes else "",
    )
    db.add(l)
    db.commit()
    db.refresh(l)
    return _lecon_dict(l)


@app.get("/api/lecons")
def list_lecons(
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    rows = (
        db.query(database.Lecon)
        .filter(database.Lecon.user_id == user.id)
        .order_by(database.Lecon.date.desc(), database.Lecon.heure.desc())
        .all()
    )
    return [_lecon_dict(l) for l in rows]


@app.get("/api/lecons/{lecon_id}")
def get_lecon(
    lecon_id: int,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    l = _get_or_404(db, database.Lecon, lecon_id)
    _check_owner(l.user_id, user.id)
    return _lecon_dict(l)


@app.patch("/api/lecons/{lecon_id}")
def patch_lecon(
    lecon_id: int,
    data: dict,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    l = _get_or_404(db, database.Lecon, lecon_id)
    _check_owner(l.user_id, user.id)
    for k, v in data.items():
        if k == "date":
            l.date = dt_date.fromisoformat(v)
        elif k == "heure":
            l.heure = dt_time.fromisoformat(v)
        elif k in {"maitre", "theme", "notes"}:
            setattr(l, k, v)
    db.commit()
    db.refresh(l)
    return _lecon_dict(l)


@app.delete("/api/lecons/{lecon_id}")
def delete_lecon(
    lecon_id: int,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    l = _get_or_404(db, database.Lecon, lecon_id)
    _check_owner(l.user_id, user.id)
    db.delete(l)
    db.commit()
    return {"ok": True}


# ─────────────────────────────────────────────────────────────────────────────
#  Compétitions
# ─────────────────────────────────────────────────────────────────────────────

class CompetitionPayload(BaseModel):
    nom: str
    date: str
    arme: Optional[str] = "epee"
    niveau: str
    ville: Optional[str] = ""
    lieu: Optional[str] = ""
    etat_de_forme: Optional[str] = ""
    a_poule: bool
    a_tableau: bool


@app.post("/api/competitions")
def create_competition(
    payload: CompetitionPayload,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    c = database.Competition(
        user_id=user.id,
        nom=payload.nom.strip(),
        date=dt_date.fromisoformat(payload.date),
        arme=payload.arme or "epee",
        niveau=payload.niveau,
        ville=payload.ville.strip() if payload.ville else "",
        lieu=payload.lieu.strip() if payload.lieu else "",
        etat_de_forme=payload.etat_de_forme.strip() if payload.etat_de_forme else "",
        a_poule=payload.a_poule,
        a_tableau=payload.a_tableau,
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return _competition_dict(c)


@app.get("/api/competitions")
def list_competitions(
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    rows = (
        db.query(database.Competition)
        .filter(database.Competition.user_id == user.id)
        .order_by(database.Competition.date.desc())
        .all()
    )
    return [_competition_dict(c) for c in rows]


@app.get("/api/competitions/{comp_id}")
def get_competition(
    comp_id: int,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    c = _get_or_404(db, database.Competition, comp_id)
    _check_owner(c.user_id, user.id)
    return _competition_dict(c)


@app.patch("/api/competitions/{comp_id}")
def patch_competition(
    comp_id: int,
    data: dict,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    c = _get_or_404(db, database.Competition, comp_id)
    _check_owner(c.user_id, user.id)
    allowed = {"terminee", "nom", "arme", "niveau", "ville", "lieu", "etat_de_forme", "notes_analyse"}
    for k, v in data.items():
        if k == "date":
            c.date = dt_date.fromisoformat(v)
        elif k in allowed:
            setattr(c, k, v)
    db.commit()
    db.refresh(c)
    return _competition_dict(c)


@app.delete("/api/competitions/{comp_id}")
def delete_competition(
    comp_id: int,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    c = _get_or_404(db, database.Competition, comp_id)
    _check_owner(c.user_id, user.id)
    # Photos : supprimer fichiers disque + BDD
    photos = db.query(database.Photo).filter(database.Photo.competition_id == comp_id).all()
    for p in photos:
        path = os.path.join(UPLOAD_DIR, p.filename)
        if os.path.exists(path):
            os.remove(path)
        db.delete(p)
    # Poule + assaults_poule
    poule = db.query(database.Poule).filter(database.Poule.competition_id == comp_id).first()
    if poule:
        db.query(database.AssaultPoule).filter(database.AssaultPoule.poule_id == poule.id).delete()
        db.delete(poule)
    # Assaults tableau
    db.query(database.AssaultTableau).filter(database.AssaultTableau.competition_id == comp_id).delete()
    db.delete(c)
    db.commit()
    return {"ok": True}


# ─────────────────────────────────────────────────────────────────────────────
#  Poules
# ─────────────────────────────────────────────────────────────────────────────

class PoulePayload(BaseModel):
    nb_tireurs: int


@app.post("/api/competitions/{comp_id}/poule")
def create_poule(
    comp_id: int,
    payload: PoulePayload,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    c = _get_or_404(db, database.Competition, comp_id)
    _check_owner(c.user_id, user.id)
    p = database.Poule(competition_id=comp_id, nb_tireurs=payload.nb_tireurs)
    db.add(p)
    db.commit()
    db.refresh(p)
    return _poule_dict(p)


@app.get("/api/competitions/{comp_id}/poule")
def get_poule(
    comp_id: int,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    c = _get_or_404(db, database.Competition, comp_id)
    _check_owner(c.user_id, user.id)
    p = db.query(database.Poule).filter(database.Poule.competition_id == comp_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Poule introuvable")
    return _poule_dict(p)


@app.patch("/api/poules/{poule_id}")
def patch_poule(
    poule_id: int,
    data: dict,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    p = _get_or_404(db, database.Poule, poule_id)
    c = _get_or_404(db, database.Competition, p.competition_id)
    _check_owner(c.user_id, user.id)
    allowed = {"qualifie", "terminee"}
    for k, v in data.items():
        if k in allowed:
            setattr(p, k, v)
    db.commit()
    db.refresh(p)
    return _poule_dict(p)


# ─────────────────────────────────────────────────────────────────────────────
#  Assaults de poule
# ─────────────────────────────────────────────────────────────────────────────

class AssaultPoulePayload(BaseModel):
    numero: int
    adversaire: Optional[str] = ""
    score_moi: Optional[int] = None
    score_adversaire: Optional[int] = None
    victoire: Optional[bool] = None
    commentaires: Optional[str] = ""


@app.post("/api/poules/{poule_id}/assaults")
def create_assault_poule(
    poule_id: int,
    payload: AssaultPoulePayload,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    p = _get_or_404(db, database.Poule, poule_id)
    c = _get_or_404(db, database.Competition, p.competition_id)
    _check_owner(c.user_id, user.id)
    a = database.AssaultPoule(
        poule_id=poule_id,
        numero=payload.numero,
        adversaire=payload.adversaire.strip() if payload.adversaire else "",
        score_moi=payload.score_moi,
        score_adversaire=payload.score_adversaire,
        victoire=payload.victoire,
        commentaires=payload.commentaires.strip() if payload.commentaires else "",
    )
    db.add(a)
    db.commit()
    db.refresh(a)
    return _assault_poule_dict(a)


@app.patch("/api/assaults_poule/{assault_id}")
def patch_assault_poule(
    assault_id: int,
    data: dict,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    a = _get_or_404(db, database.AssaultPoule, assault_id)
    p = _get_or_404(db, database.Poule, a.poule_id)
    c = _get_or_404(db, database.Competition, p.competition_id)
    _check_owner(c.user_id, user.id)
    allowed = {"adversaire", "score_moi", "score_adversaire", "victoire", "commentaires", "notes_post"}
    for k, v in data.items():
        if k in allowed:
            setattr(a, k, v)
    db.commit()
    db.refresh(a)
    return _assault_poule_dict(a)


@app.delete("/api/assaults_poule/{assault_id}")
def delete_assault_poule(
    assault_id: int,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    a = _get_or_404(db, database.AssaultPoule, assault_id)
    p = _get_or_404(db, database.Poule, a.poule_id)
    c = _get_or_404(db, database.Competition, p.competition_id)
    _check_owner(c.user_id, user.id)
    db.delete(a)
    db.commit()
    return {"ok": True}


@app.get("/api/poules/{poule_id}/assaults")
def list_assaults_poule(
    poule_id: int,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    p = _get_or_404(db, database.Poule, poule_id)
    c = _get_or_404(db, database.Competition, p.competition_id)
    _check_owner(c.user_id, user.id)
    rows = (
        db.query(database.AssaultPoule)
        .filter(database.AssaultPoule.poule_id == poule_id)
        .order_by(database.AssaultPoule.numero)
        .all()
    )
    return [_assault_poule_dict(a) for a in rows]


# ─────────────────────────────────────────────────────────────────────────────
#  Assaults de tableau
# ─────────────────────────────────────────────────────────────────────────────

class AssaultTableauPayload(BaseModel):
    tour: int
    adversaire: Optional[str] = ""
    score_moi: Optional[int] = None
    score_adversaire: Optional[int] = None
    victoire: bool
    commentaires: Optional[str] = ""


@app.post("/api/competitions/{comp_id}/tableau")
def create_assault_tableau(
    comp_id: int,
    payload: AssaultTableauPayload,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    c = _get_or_404(db, database.Competition, comp_id)
    _check_owner(c.user_id, user.id)
    a = database.AssaultTableau(
        competition_id=comp_id,
        tour=payload.tour,
        adversaire=payload.adversaire.strip() if payload.adversaire else "",
        score_moi=payload.score_moi,
        score_adversaire=payload.score_adversaire,
        victoire=payload.victoire,
        commentaires=payload.commentaires.strip() if payload.commentaires else "",
    )
    db.add(a)
    db.commit()
    db.refresh(a)
    return _assault_tableau_dict(a)


@app.patch("/api/assaults_tableau/{assault_id}")
def patch_assault_tableau(
    assault_id: int,
    data: dict,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    a = _get_or_404(db, database.AssaultTableau, assault_id)
    c = _get_or_404(db, database.Competition, a.competition_id)
    _check_owner(c.user_id, user.id)
    allowed = {"adversaire", "score_moi", "score_adversaire", "victoire", "commentaires", "notes_post"}
    for k, v in data.items():
        if k in allowed:
            setattr(a, k, v)
    db.commit()
    db.refresh(a)
    return _assault_tableau_dict(a)


@app.delete("/api/assaults_tableau/{assault_id}")
def delete_assault_tableau(
    assault_id: int,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    a = _get_or_404(db, database.AssaultTableau, assault_id)
    c = _get_or_404(db, database.Competition, a.competition_id)
    _check_owner(c.user_id, user.id)
    db.delete(a)
    db.commit()
    return {"ok": True}


@app.get("/api/competitions/{comp_id}/tableau")
def list_assaults_tableau(
    comp_id: int,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    c = _get_or_404(db, database.Competition, comp_id)
    _check_owner(c.user_id, user.id)
    rows = (
        db.query(database.AssaultTableau)
        .filter(database.AssaultTableau.competition_id == comp_id)
        .order_by(database.AssaultTableau.created_at)
        .all()
    )
    return [_assault_tableau_dict(a) for a in rows]


# ─────────────────────────────────────────────────────────────────────────────
#  Export XLSX
# ─────────────────────────────────────────────────────────────────────────────

def _tour_label_py(n: int) -> str:
    if n == 2:  return "Finale"
    if n == 4:  return "Demi-finale"
    if n == 8:  return "Quart de finale"
    return f"Tour de {n}"


def _xlsx_competitions(user_id: int, db: Session) -> io.BytesIO:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Compétitions"
    ws1.append([
        "ID", "Nom", "Date", "Arme", "Niveau", "Ville", "Lieu",
        "Etat_de_forme", "A_poule", "A_tableau", "Terminée", "Notes_analyse",
    ])
    comps = (
        db.query(database.Competition)
        .filter(database.Competition.user_id == user_id)
        .order_by(database.Competition.date)
        .all()
    )
    for c in comps:
        ws1.append([
            c.id, c.nom, str(c.date), c.arme, c.niveau,
            c.ville or "", c.lieu or "", c.etat_de_forme or "",
            c.a_poule, c.a_tableau, c.terminee, c.notes_analyse or "",
        ])

    ws2 = wb.create_sheet("Assaults_Poule")
    ws2.append([
        "Competition_ID", "Competition_Nom", "Competition_Date",
        "Poule_ID", "Numero", "Adversaire",
        "Score_moi", "Score_adversaire", "Victoire",
        "Commentaires", "Notes_post",
    ])
    for c in comps:
        poule = db.query(database.Poule).filter(database.Poule.competition_id == c.id).first()
        if not poule:
            continue
        for a in (
            db.query(database.AssaultPoule)
            .filter(database.AssaultPoule.poule_id == poule.id)
            .order_by(database.AssaultPoule.numero)
            .all()
        ):
            ws2.append([
                c.id, c.nom, str(c.date),
                poule.id, a.numero, a.adversaire or "",
                a.score_moi, a.score_adversaire, a.victoire,
                a.commentaires or "", a.notes_post or "",
            ])

    ws3 = wb.create_sheet("Assaults_Tableau")
    ws3.append([
        "Competition_ID", "Competition_Nom", "Competition_Date",
        "Tour_label", "Tour_num", "Adversaire",
        "Score_moi", "Score_adversaire", "Victoire",
        "Commentaires", "Notes_post",
    ])
    for c in comps:
        for a in (
            db.query(database.AssaultTableau)
            .filter(database.AssaultTableau.competition_id == c.id)
            .order_by(database.AssaultTableau.created_at)
            .all()
        ):
            ws3.append([
                c.id, c.nom, str(c.date),
                _tour_label_py(a.tour), a.tour, a.adversaire or "",
                a.score_moi, a.score_adversaire, a.victoire,
                a.commentaires or "", a.notes_post or "",
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _xlsx_entrainement(user_id: int, db: Session) -> io.BytesIO:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Assaults_libres"
    ws1.append(["ID", "Date", "Heure", "Notes"])
    for a in (
        db.query(database.Assault)
        .filter(database.Assault.user_id == user_id)
        .order_by(database.Assault.date, database.Assault.heure)
        .all()
    ):
        ws1.append([a.id, str(a.date), str(a.heure), a.notes or ""])

    ws2 = wb.create_sheet("Leçons")
    ws2.append(["ID", "Date", "Heure", "Maitre", "Theme", "Notes"])
    for l in (
        db.query(database.Lecon)
        .filter(database.Lecon.user_id == user_id)
        .order_by(database.Lecon.date, database.Lecon.heure)
        .all()
    ):
        ws2.append([l.id, str(l.date), str(l.heure), l.maitre or "", l.theme or "", l.notes or ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@app.get("/api/export/competitions")
def export_competitions_xlsx(
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    buf = _xlsx_competitions(user.id, db)
    return StreamingResponse(buf, media_type=XLSX_MIME,
        headers={"Content-Disposition": "attachment; filename=med_competitions.xlsx"})


@app.get("/api/export/entrainement")
def export_entrainement_xlsx(
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    buf = _xlsx_entrainement(user.id, db)
    return StreamingResponse(buf, media_type=XLSX_MIME,
        headers={"Content-Disposition": "attachment; filename=med_entrainement.xlsx"})


@app.get("/api/admin/users/{user_id}/export/competitions")
def admin_export_competitions(
    user_id: int,
    request: Request,
    db: Session = Depends(database.get_db),
):
    require_admin(request)
    user = db.query(database.User).filter(database.User.id == user_id).first()
    name = (user.name if user else str(user_id)).replace(" ", "_")
    buf = _xlsx_competitions(user_id, db)
    return StreamingResponse(buf, media_type=XLSX_MIME,
        headers={"Content-Disposition": f"attachment; filename=med_competitions_{name}.xlsx"})


@app.get("/api/admin/users/{user_id}/export/entrainement")
def admin_export_entrainement(
    user_id: int,
    request: Request,
    db: Session = Depends(database.get_db),
):
    require_admin(request)
    user = db.query(database.User).filter(database.User.id == user_id).first()
    name = (user.name if user else str(user_id)).replace(" ", "_")
    buf = _xlsx_entrainement(user_id, db)
    return StreamingResponse(buf, media_type=XLSX_MIME,
        headers={"Content-Disposition": f"attachment; filename=med_entrainement_{name}.xlsx"})


# ─────────────────────────────────────────────────────────────────────────────
#  Photos
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/competitions/{comp_id}/photos")
async def upload_photo(
    comp_id: int,
    type_photo: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    c = _get_or_404(db, database.Competition, comp_id)
    _check_owner(c.user_id, user.id)
    ext = os.path.splitext(file.filename)[1].lower() or ".jpg"
    filename = f"{comp_id}_{type_photo}_{uuid.uuid4().hex[:8]}{ext}"
    path = os.path.join(UPLOAD_DIR, filename)
    with open(path, "wb") as f:
        f.write(await file.read())
    photo = database.Photo(competition_id=comp_id, type_photo=type_photo, filename=filename)
    db.add(photo)
    db.commit()
    db.refresh(photo)
    return {"id": photo.id, "filename": filename, "url": f"/static/uploads/{filename}"}


@app.get("/api/competitions/{comp_id}/photos")
def list_photos(
    comp_id: int,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    c = _get_or_404(db, database.Competition, comp_id)
    _check_owner(c.user_id, user.id)
    rows = db.query(database.Photo).filter(database.Photo.competition_id == comp_id).all()
    return [
        {"id": p.id, "type_photo": p.type_photo, "filename": p.filename,
         "url": f"/static/uploads/{p.filename}"}
        for p in rows
    ]


@app.delete("/api/photos/{photo_id}")
def delete_photo(
    photo_id: int,
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    p = _get_or_404(db, database.Photo, photo_id)
    c = _get_or_404(db, database.Competition, p.competition_id)
    _check_owner(c.user_id, user.id)
    path = os.path.join(UPLOAD_DIR, p.filename)
    if os.path.exists(path):
        os.remove(path)
    db.delete(p)
    db.commit()
    return {"ok": True}


# ─────────────────────────────────────────────────────────────────────────────
#  Export Dashboard HTML
# ─────────────────────────────────────────────────────────────────────────────

_EI_URL = (
    "https://www.escrime-info.com/gregxml/Greg/BD/services/service.php/tireur?id=32724"
)

def _safe_json(obj) -> str:
    return json.dumps(obj, ensure_ascii=False).replace("</script>", "<\\/script>")


_DASHBOARD_TEMPLATE = """\
<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>David Coldefy · Escrime</title>
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f0f2f5;color:#1a1a2e}
header{background:#1e3a5f;color:#fff;padding:1.25rem 2rem;display:flex;justify-content:space-between;align-items:flex-end;flex-wrap:wrap;gap:.5rem}
header h1{font-size:1.5rem;font-weight:700}
.header-sub{font-size:.8rem;color:#94b4d1;margin-top:.2rem}
.export-date{font-size:.75rem;color:#7a9ec2;white-space:nowrap}
.stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:.75rem;padding:1.25rem 2rem}
.stat{background:#fff;border-radius:8px;padding:.9rem 1rem;box-shadow:0 1px 3px rgba(0,0,0,.09)}
.stat-value{font-size:1.7rem;font-weight:700;color:#1e3a5f;line-height:1}
.stat-label{font-size:.7rem;color:#888;margin-top:.3rem;text-transform:uppercase;letter-spacing:.05em}
.tabs{display:flex;padding:0 2rem;background:#fff;border-bottom:2px solid #e5e7eb}
.tab{padding:.7rem 1.2rem;cursor:pointer;font-weight:500;color:#777;border-bottom:3px solid transparent;margin-bottom:-2px;font-size:.9rem;transition:color .15s}
.tab.active{color:#1e3a5f;border-color:#1e3a5f}
.tab:hover:not(.active){color:#333}
.content{padding:1.25rem 2rem}
.panel{display:none}.panel.active{display:block}
.card{background:#fff;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.09);overflow:hidden}
.card+.card{margin-top:.75rem}
.card-head{padding:.75rem 1rem;font-size:.75rem;font-weight:600;color:#888;text-transform:uppercase;letter-spacing:.05em;border-bottom:1px solid #f3f4f6}
.chart-row{display:grid;grid-template-columns:2fr 1fr;gap:.75rem;margin-bottom:1rem}
.chart-box{padding:1rem}
.filters{display:flex;gap:.6rem;margin-bottom:.75rem;flex-wrap:wrap;align-items:center}
.filters input,.filters select{padding:.45rem .7rem;border:1px solid #d1d5db;border-radius:6px;font-size:.85rem;background:#fff;color:#374151}
.filters input{min-width:180px}
.count{font-size:.8rem;color:#888;margin-left:auto}
.tbl-wrap{overflow-x:auto}
table{width:100%;border-collapse:collapse;font-size:.85rem}
thead th{background:#f8fafc;padding:.65rem 1rem;text-align:left;font-weight:600;color:#374151;cursor:pointer;user-select:none;white-space:nowrap;border-bottom:1px solid #e5e7eb}
thead th:hover{background:#f0f2f5}
thead th.asc::after{content:' ↑'}
thead th.desc::after{content:' ↓'}
tbody tr{border-bottom:1px solid #f3f4f6}
tbody tr:hover{background:#f8fafc}
tbody td{padding:.55rem 1rem;color:#374151;vertical-align:middle}
.rank{display:inline-block;padding:.12rem .45rem;border-radius:4px;font-weight:700;font-size:.8rem}
.r1{background:#fef3c7;color:#92400e}
.r3{background:#f1f5f9;color:#374151}
.r8{background:#d1fae5;color:#065f46}
.r16{background:#dbeafe;color:#1e40af}
.rn{background:#f3f4f6;color:#9ca3af}
.cat-s{background:#dbeafe;color:#1e40af;padding:.1rem .4rem;border-radius:3px;font-size:.72rem;font-weight:700}
.cat-v{background:#fed7aa;color:#c2410c;padding:.1rem .4rem;border-radius:3px;font-size:.72rem;font-weight:700}
.comp-card{background:#fff;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.09);margin-bottom:.75rem;overflow:hidden}
.comp-hd{padding:.9rem 1.25rem;cursor:pointer;display:flex;justify-content:space-between;align-items:center;gap:.5rem}
.comp-hd:hover{background:#f8fafc}
.comp-title{font-weight:600;font-size:.95rem}
.comp-meta{font-size:.78rem;color:#888;margin-top:.15rem}
.badge-ok{background:#d1fae5;color:#065f46;padding:.2rem .65rem;border-radius:20px;font-size:.72rem;font-weight:700}
.badge-wip{background:#fef3c7;color:#92400e;padding:.2rem .65rem;border-radius:20px;font-size:.72rem;font-weight:700}
.comp-body{display:none;padding:.25rem 1.25rem 1rem}
.comp-body.open{display:block}
.section-title{font-size:.72rem;font-weight:700;color:#888;text-transform:uppercase;letter-spacing:.05em;padding:.6rem 0 .35rem;border-bottom:1px solid #f3f4f6;margin-top:.6rem}
.v{color:#16a34a;font-weight:700}.d{color:#dc2626;font-weight:700}.score{font-weight:600}
.etat{background:#f0f9ff;padding:.65rem .85rem;border-radius:6px;font-style:italic;font-size:.85rem;color:#374151;white-space:pre-wrap;line-height:1.5}
.lecon{background:#fff;border-radius:6px;padding:.7rem 1rem;margin-bottom:.5rem;box-shadow:0 1px 2px rgba(0,0,0,.07)}
.lecon-hd{display:flex;gap:.75rem;align-items:baseline;margin-bottom:.25rem}
.lecon-date{font-size:.75rem;color:#888}
.lecon-maitre{font-size:.78rem;font-weight:600;color:#1e3a5f}
.lecon-theme{font-size:.85rem;font-weight:500}
.lecon-notes{font-size:.82rem;color:#555;white-space:pre-wrap;line-height:1.5;margin-top:.25rem}
.section-h3{font-size:.9rem;font-weight:600;color:#374151;margin:1rem 0 .5rem}
.no-ffe{background:#fff3cd;border:1px solid #ffc107;border-radius:6px;padding:.75rem 1rem;color:#856404;font-size:.85rem;margin-bottom:1rem}
@media(max-width:640px){.stats,.content{padding:1rem}.chart-row{grid-template-columns:1fr}header{padding:1rem}}
</style>
</head>
<body>
<header>
  <div><h1>David Coldefy &middot; &Eacute;p&eacute;e</h1><div class="header-sub" id="hdr-sub"></div></div>
  <div class="export-date" id="hdr-date"></div>
</header>
<div class="stats" id="stats-grid"></div>
<div class="tabs">
  <div class="tab active" data-tab="palmares">Palmar&egrave;s FFE</div>
  <div class="tab" data-tab="med">Comp&eacute;titions MED</div>
  <div class="tab" data-tab="training">Entra&icirc;nement</div>
</div>
<div class="content">
  <div class="panel active" id="panel-palmares">
    <div class="chart-row">
      <div class="card chart-box"><div class="card-head">&Eacute;volution des classements</div><canvas id="chart-evo" style="max-height:280px"></canvas></div>
      <div class="card chart-box"><div class="card-head">Comp&eacute;titions par ann&eacute;e</div><canvas id="chart-year" style="max-height:280px"></canvas></div>
    </div>
    <div id="ffe-warning"></div>
    <div class="filters">
      <input type="text" id="search" placeholder="Rechercher...">
      <select id="f-year"><option value="">Toutes les ann&eacute;es</option></select>
      <select id="f-cat"><option value="">Toutes cat&eacute;gories</option><option value="S">Senior</option><option value="V">V&eacute;t&eacute;ran</option></select>
      <select id="f-equipe"><option value="">Ind. + &Eacute;quipe</option><option value="0">Individuel</option><option value="1">&Eacute;quipe</option></select>
      <span class="count" id="count"></span>
    </div>
    <div class="card tbl-wrap">
      <table id="tbl">
        <thead><tr>
          <th data-k="date">Date</th><th data-k="libelle">Comp&eacute;tition</th>
          <th data-k="lieu">Lieu</th><th data-k="categorie">Cat.</th>
          <th data-k="equipe">&Eacute;q.</th><th data-k="classement">Classement</th>
        </tr></thead>
        <tbody id="tbody"></tbody>
      </table>
    </div>
  </div>
  <div class="panel" id="panel-med"><div id="med-list"></div></div>
  <div class="panel" id="panel-training"><div id="training-list"></div></div>
</div>
<script>
__DATA_JS__
</script>
<script>
function fd(s){if(!s)return'';var p=s.slice(0,10).split('-');return p[2]+'/'+p[1]+'/'+p[0]}
function esc(s){return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;')}
function rankCls(r){if(!r)return 'rn';if(r===1)return 'r1';if(r<=3)return 'r3';if(r<=8)return 'r8';if(r<=16)return 'r16';return 'rn'}
// Onglets — initialisés en premier, sans dépendance externe
document.querySelectorAll('.tab').forEach(function(tab){tab.addEventListener('click',function(){document.querySelectorAll('.tab').forEach(function(t){t.classList.remove('active')});document.querySelectorAll('.panel').forEach(function(p){p.classList.remove('active')});tab.classList.add('active');document.getElementById('panel-'+tab.dataset.tab).classList.add('active');});});
(function(){
  var nb=DATA.palmares.length;
  var years=DATA.palmares.map(function(c){return c.date?c.date.slice(0,4):null}).filter(Boolean);
  var sub=nb?nb+' compétitions FFE · '+(Math.min.apply(null,years))+'–'+(Math.max.apply(null,years)):'Palmarès FFE non disponible';
  document.getElementById('hdr-sub').textContent=sub;
  var d=new Date(DATA.exported_at);
  document.getElementById('hdr-date').textContent='Export du '+d.toLocaleDateString('fr-FR');
})();
(function(){
  var comps=DATA.palmares;
  var ranked=comps.filter(function(c){return c.classement});
  var seniors=comps.filter(function(c){return c.categorie==='S'});
  var vets=comps.filter(function(c){return c.categorie==='V'});
  function best(arr){return arr.length?Math.min.apply(null,arr.map(function(c){return c.classement})):null}
  var cards=[
    {v:comps.length||'—',l:'Compétitions FFE'},
    {v:best(ranked)?'#'+best(ranked):'—',l:'Meilleur rang'},
    {v:seniors.length,l:'Senior'},{v:best(ranked&&seniors)?'#'+best(seniors.filter(function(c){return c.classement})):'—',l:'Meilleur rang S.'},
    {v:vets.length,l:'Vétéran'},{v:best(vets.filter(function(c){return c.classement}))?'#'+best(vets.filter(function(c){return c.classement})):'—',l:'Meilleur rang Vét.'},
    {v:DATA.med_competitions.length,l:'Compétitions MED'},{v:DATA.med_lecons.length,l:'Leçons MED'},
  ];
  document.getElementById('stats-grid').innerHTML=cards.map(function(c){
    return'<div class="stat"><div class="stat-value">'+c.v+'</div><div class="stat-label">'+c.l+'</div></div>'
  }).join('');
})();
(function(){
  if(!DATA.palmares.length){document.getElementById('ffe-warning').innerHTML='<div class="no-ffe">Palmarès FFE non disponible lors de cet export (escrime-info.com inaccessible).</div>';return;}
  var s=document.createElement('script');
  s.src='https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js';
  s.onerror=function(){document.getElementById('ffe-warning').innerHTML='<div class="no-ffe">Graphiques indisponibles (Chart.js non chargé — vérifier la connexion internet).</div>';};
  s.onload=function(){
    try{
      var comps=DATA.palmares.filter(function(c){return c.classement&&c.date});
      function toDs(cat){return comps.filter(function(c){return c.categorie===cat}).map(function(c){var yr=parseInt(c.date.slice(0,4)),mo=parseInt(c.date.slice(5,7));return{x:yr+(mo-1)/12,y:c.classement,label:c.libelle,lieu:c.lieu||'',date:fd(c.date)}});}
      new Chart(document.getElementById('chart-evo'),{type:'scatter',data:{datasets:[
        {label:'Senior',data:toDs('S'),backgroundColor:'rgba(59,130,246,.65)',pointRadius:5,pointHoverRadius:7},
        {label:'Vétéran',data:toDs('V'),backgroundColor:'rgba(249,115,22,.65)',pointRadius:5,pointHoverRadius:7},
      ]},options:{responsive:true,plugins:{legend:{position:'top'},tooltip:{callbacks:{label:function(ctx){var d=ctx.raw;return[d.label,d.lieu+' · '+d.date,'Classement : '+d.y]}}}},
      scales:{x:{title:{display:true,text:'Année'},min:2007,max:2027,ticks:{stepSize:1,callback:function(v){return Number.isInteger(v)?v:''}}},y:{reverse:true,title:{display:true,text:'Classement'},min:1,ticks:{stepSize:10}}}}});
      var years={};DATA.palmares.forEach(function(c){var y=c.date.slice(0,4);years[y]=(years[y]||0)+1});
      var yl=Object.keys(years).sort();
      new Chart(document.getElementById('chart-year'),{type:'bar',data:{labels:yl,datasets:[{label:'Compétitions',data:yl.map(function(y){return years[y]}),backgroundColor:'#1e3a5f',borderRadius:3}]},options:{responsive:true,plugins:{legend:{display:false}},scales:{y:{beginAtZero:true,ticks:{stepSize:1}},x:{ticks:{maxRotation:90}}}}});
    }catch(e){console.error('Chart error:',e);}
  };
  document.head.appendChild(s);
})();
var sortK='date',sortD=-1,filtered=DATA.palmares.slice();
function sortArr(arr,k,d){return arr.slice().sort(function(a,b){var va=a[k]!=null?a[k]:'',vb=b[k]!=null?b[k]:'';if(k==='classement'){va=va||9999;vb=vb||9999;}return va<vb?-d:va>vb?d:0});}
function applyFilters(){var q=(document.getElementById('search').value||'').toLowerCase(),yr=document.getElementById('f-year').value,cat=document.getElementById('f-cat').value,eq=document.getElementById('f-equipe').value;filtered=DATA.palmares.filter(function(c){if(q&&!(c.libelle||'').toLowerCase().includes(q)&&!(c.lieu||'').toLowerCase().includes(q))return false;if(yr&&!c.date.startsWith(yr))return false;if(cat&&c.categorie!==cat)return false;if(eq!==''&&String(c.equipe?1:0)!==eq)return false;return true;});filtered=sortArr(filtered,sortK,sortD);renderTable();}
function renderTable(){document.getElementById('tbody').innerHTML=filtered.map(function(c){var rk=c.classement?'<span class="rank '+rankCls(c.classement)+'">'+c.classement+'</span>':'—';var cat=c.categorie?'<span class="cat-'+c.categorie.toLowerCase()+'">'+c.categorie+'</span>':'';return'<tr><td>'+fd(c.date)+'</td><td>'+esc(c.libelle)+'</td><td>'+esc(c.lieu||'')+'</td><td>'+cat+'</td><td>'+(c.equipe?'E':'I')+'</td><td>'+rk+'</td></tr>';}).join('');document.getElementById('count').textContent=filtered.length+' / '+DATA.palmares.length+' compétitions';}
(function(){var years=[...new Set(DATA.palmares.map(function(c){return c.date.slice(0,4)}))].sort().reverse();var sel=document.getElementById('f-year');years.forEach(function(y){var o=document.createElement('option');o.value=y;o.textContent=y;sel.appendChild(o)});})();
document.querySelectorAll('#tbl thead th[data-k]').forEach(function(th){th.addEventListener('click',function(){var k=th.dataset.k;if(sortK===k)sortD=-sortD;else{sortK=k;sortD=k==='date'?-1:1;}document.querySelectorAll('#tbl thead th').forEach(function(t){t.classList.remove('asc','desc')});th.classList.add(sortD===1?'asc':'desc');filtered=sortArr(filtered,sortK,sortD);renderTable();});});
['search','f-year','f-cat','f-equipe'].forEach(function(id){document.getElementById(id).addEventListener('input',applyFilters);document.getElementById(id).addEventListener('change',applyFilters);});
filtered=sortArr(DATA.palmares,'date',-1);renderTable();
(function(){var el=document.getElementById('med-list');if(!DATA.med_competitions.length){el.innerHTML='<p style="color:#888;padding:1rem">Aucune compétition MED.</p>';return;}
el.innerHTML=DATA.med_competitions.map(function(c){
  var pouleH='';
  if(c.poule){var qual=c.poule.qualifie===true?'✓ Qualifié':c.poule.qualifie===false?'✗ Non qualifié':'';pouleH='<div class="section-title">Poule &mdash; '+c.poule.nb_tireurs+' tireurs'+(qual?' &middot; '+qual:'')+'</div><table style="width:100%;font-size:.82rem"><thead><tr style="color:#888"><th style="text-align:left;padding:.3rem .5rem">Adversaire</th><th style="padding:.3rem .5rem">Score</th><th style="padding:.3rem .5rem">Rés.</th><th style="text-align:left;padding:.3rem .5rem">Notes</th></tr></thead><tbody>'+(c.poule.assaults||[]).map(function(a){return'<tr style="border-bottom:1px solid #f3f4f6"><td style="padding:.3rem .5rem">'+esc(a.adversaire||'—')+'</td><td class="score" style="text-align:center;padding:.3rem .5rem">'+(a.score_moi!=null?a.score_moi:'')+'–'+(a.score_adversaire!=null?a.score_adversaire:'')+'</td><td style="text-align:center;padding:.3rem .5rem"><span class="'+(a.victoire?'v':'d')+'">'+(a.victoire?'V':'D')+'</span></td><td style="color:#666;padding:.3rem .5rem;font-size:.78rem">'+esc(a.notes_post||a.commentaires||'')+'</td></tr>';}).join('')+'</tbody></table>';}
  var tableauH='';
  if(c.tableau&&c.tableau.length){tableauH='<div class="section-title">Tableau</div><table style="width:100%;font-size:.82rem"><thead><tr style="color:#888"><th style="text-align:left;padding:.3rem .5rem">Tour</th><th style="text-align:left;padding:.3rem .5rem">Adversaire</th><th style="padding:.3rem .5rem">Score</th><th style="padding:.3rem .5rem">Rés.</th><th style="text-align:left;padding:.3rem .5rem">Notes</th></tr></thead><tbody>'+c.tableau.map(function(a){return'<tr style="border-bottom:1px solid #f3f4f6"><td style="padding:.3rem .5rem;font-weight:500">'+esc(a.tour_label)+'</td><td style="padding:.3rem .5rem">'+esc(a.adversaire||'—')+'</td><td class="score" style="text-align:center;padding:.3rem .5rem">'+(a.score_moi!=null?a.score_moi:'')+'–'+(a.score_adversaire!=null?a.score_adversaire:'')+'</td><td style="text-align:center;padding:.3rem .5rem"><span class="'+(a.victoire?'v':'d')+'">'+(a.victoire?'V':'D')+'</span></td><td style="color:#666;padding:.3rem .5rem;font-size:.78rem">'+esc(a.notes_post||a.commentaires||'')+'</td></tr>';}).join('')+'</tbody></table>';}
  var etH=c.etat_de_forme?'<div class="section-title">État de forme</div><div class="etat">'+esc(c.etat_de_forme)+'</div>':'';
  var naH=c.notes_analyse?'<div class="section-title">Analyse</div><div class="etat">'+esc(c.notes_analyse)+'</div>':'';
  return'<div class="comp-card"><div class="comp-hd" onclick="this.nextElementSibling.classList.toggle(\'open\')"><div><div class="comp-title">'+esc(c.nom)+'</div><div class="comp-meta">'+fd(c.date)+(c.ville?' · '+esc(c.ville):'')+(c.niveau?' · '+esc(c.niveau):'')+'</div></div><span class="'+(c.terminee?'badge-ok':'badge-wip')+'">'+(c.terminee?'Terminée':'En cours')+'</span></div><div class="comp-body">'+etH+pouleH+tableauH+naH+'</div></div>';
}).join('');})();
(function(){var el=document.getElementById('training-list');var lecons=DATA.med_lecons||[];var assaults=DATA.med_assaults||[];if(!lecons.length&&!assaults.length){el.innerHTML='<p style="color:#888;padding:1rem">Aucune donnée d\'entraînement.</p>';return;}
var html='';
if(lecons.length){html+='<div class="section-h3">Leçons ('+lecons.length+')</div>';html+=lecons.map(function(l){return'<div class="lecon"><div class="lecon-hd"><span class="lecon-date">'+fd(l.date)+(l.heure?' · '+l.heure.slice(0,5):'')+'</span>'+(l.maitre?'<span class="lecon-maitre">'+esc(l.maitre)+'</span>':'')+'</div>'+(l.theme?'<div class="lecon-theme">'+esc(l.theme)+'</div>':'')+(l.notes?'<div class="lecon-notes">'+esc(l.notes)+'</div>':'')+'</div>';}).join('');}
if(assaults.length){html+='<div class="section-h3">Assaults libres ('+assaults.length+')</div>';html+=assaults.map(function(a){return'<div class="lecon"><div class="lecon-date">'+fd(a.date)+(a.heure?' · '+a.heure.slice(0,5):'')+'</div>'+(a.notes?'<div class="lecon-notes">'+esc(a.notes)+'</div>':'')+'</div>';}).join('');}
el.innerHTML=html;})();
</script>
</body>
</html>
"""


def _build_dashboard_html(user_id: int, db: Session) -> str:
    now = datetime.now().isoformat(timespec="seconds")

    # ── MED competitions ──────────────────────────────────────────────────────
    comps_db = (
        db.query(database.Competition)
        .filter(database.Competition.user_id == user_id)
        .order_by(database.Competition.date.desc())
        .all()
    )
    competitions = []
    for c in comps_db:
        comp = {
            "id": c.id, "nom": c.nom, "date": str(c.date),
            "arme": c.arme, "niveau": c.niveau,
            "ville": c.ville or "", "lieu": c.lieu or "",
            "etat_de_forme": c.etat_de_forme or "",
            "a_poule": c.a_poule, "a_tableau": c.a_tableau,
            "terminee": c.terminee, "notes_analyse": c.notes_analyse or "",
            "poule": None, "tableau": [],
        }
        poule = db.query(database.Poule).filter(database.Poule.competition_id == c.id).first()
        if poule:
            ap = (db.query(database.AssaultPoule)
                  .filter(database.AssaultPoule.poule_id == poule.id)
                  .order_by(database.AssaultPoule.numero).all())
            comp["poule"] = {
                "nb_tireurs": poule.nb_tireurs, "qualifie": poule.qualifie,
                "terminee": poule.terminee,
                "assaults": [{
                    "numero": a.numero, "adversaire": a.adversaire or "",
                    "score_moi": a.score_moi, "score_adversaire": a.score_adversaire,
                    "victoire": a.victoire,
                    "commentaires": a.commentaires or "", "notes_post": a.notes_post or "",
                } for a in ap],
            }
        at = (db.query(database.AssaultTableau)
              .filter(database.AssaultTableau.competition_id == c.id)
              .order_by(database.AssaultTableau.created_at).all())
        comp["tableau"] = [{
            "tour": a.tour, "tour_label": _tour_label_py(a.tour),
            "adversaire": a.adversaire or "",
            "score_moi": a.score_moi, "score_adversaire": a.score_adversaire,
            "victoire": a.victoire,
            "commentaires": a.commentaires or "", "notes_post": a.notes_post or "",
        } for a in at]
        competitions.append(comp)

    # ── MED training ─────────────────────────────────────────────────────────
    assaults_db = (db.query(database.Assault)
                   .filter(database.Assault.user_id == user_id)
                   .order_by(database.Assault.date.desc(), database.Assault.heure.desc()).all())
    lecons_db = (db.query(database.Lecon)
                 .filter(database.Lecon.user_id == user_id)
                 .order_by(database.Lecon.date.desc(), database.Lecon.heure.desc()).all())

    # ── Escrime-info.com ─────────────────────────────────────────────────────
    ei_comps, ei_tireur = [], {}
    try:
        req = _urllib_req.Request(
            _EI_URL, headers={"User-Agent": "Mozilla/5.0 (compatible; MED/1.0)"},
        )
        with _urllib_req.urlopen(req, timeout=10) as resp:
            raw = json.loads(resp.read().decode("utf-8"))
        ei_comps = raw.get("competitions", [])
        ei_tireur = {k: raw.get(k) for k in ("id", "nom", "prenom", "naissance", "genre")}
    except Exception:
        pass  # continue without FFE data if unreachable

    data_js = "const DATA = " + _safe_json({
        "palmares": ei_comps,
        "tireur": ei_tireur,
        "med_competitions": competitions,
        "med_assaults": [_assault_dict(a) for a in assaults_db],
        "med_lecons": [_lecon_dict(l) for l in lecons_db],
        "exported_at": now,
    }) + ";"

    return _DASHBOARD_TEMPLATE.replace("__DATA_JS__", data_js)


@app.get("/api/export/dashboard")
def export_dashboard(
    db: Session = Depends(database.get_db),
    user: database.User = Depends(get_current_user),
):
    html = _build_dashboard_html(user.id, db)
    return Response(
        content=html.encode("utf-8"),
        media_type="text/html; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="palmares_david_coldefy.html"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _get_or_404(db, model, pk):
    obj = db.query(model).filter(model.id == pk).first()
    if not obj:
        raise HTTPException(status_code=404, detail=f"{model.__name__} introuvable")
    return obj


def _check_owner(resource_user_id: int, current_user_id: int):
    if resource_user_id != current_user_id:
        raise HTTPException(status_code=403, detail="Accès refusé")


def _user_dict(u: database.User) -> dict:
    return {"id": u.id, "name": u.name, "token": u.token,
            "created_at": u.created_at.isoformat() if u.created_at else None}


def _lecon_dict(l: database.Lecon) -> dict:
    return {
        "id": l.id, "date": str(l.date), "heure": str(l.heure),
        "maitre": l.maitre or "", "theme": l.theme or "", "notes": l.notes or "",
    }


def _assault_dict(a: database.Assault) -> dict:
    return {
        "id": a.id, "date": str(a.date), "heure": str(a.heure),
        "type_seance": a.type_seance, "notes": a.notes,
    }


def _competition_dict(c: database.Competition) -> dict:
    return {
        "id": c.id, "nom": c.nom, "date": str(c.date), "arme": c.arme,
        "niveau": c.niveau, "ville": c.ville, "lieu": c.lieu,
        "etat_de_forme": c.etat_de_forme, "a_poule": c.a_poule,
        "a_tableau": c.a_tableau, "terminee": c.terminee,
        "notes_analyse": c.notes_analyse or "",
        "created_at": c.created_at.isoformat() if c.created_at else None,
    }


def _poule_dict(p: database.Poule) -> dict:
    return {
        "id": p.id, "competition_id": p.competition_id,
        "nb_tireurs": p.nb_tireurs, "nb_assaults": p.nb_tireurs - 1,
        "qualifie": p.qualifie, "terminee": p.terminee,
    }


def _assault_poule_dict(a: database.AssaultPoule) -> dict:
    return {
        "id": a.id, "poule_id": a.poule_id, "numero": a.numero,
        "adversaire": a.adversaire, "score_moi": a.score_moi,
        "score_adversaire": a.score_adversaire, "victoire": a.victoire,
        "commentaires": a.commentaires, "notes_post": a.notes_post or "",
    }


def _assault_tableau_dict(a: database.AssaultTableau) -> dict:
    return {
        "id": a.id, "competition_id": a.competition_id, "tour": a.tour,
        "adversaire": a.adversaire, "score_moi": a.score_moi,
        "score_adversaire": a.score_adversaire, "victoire": a.victoire,
        "commentaires": a.commentaires, "notes_post": a.notes_post or "",
    }
